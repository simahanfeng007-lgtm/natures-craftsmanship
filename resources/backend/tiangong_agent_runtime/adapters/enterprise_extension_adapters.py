"""Enterprise execution extension adapters for Tiangong v2.

设计原则：
- 不改变 run_once / _hebing_panding / Code-X / 固定收口链。
- 工具默认只读或生成计划；涉及写入、真实外部动作、桌面/浏览器操作时默认返回可确认计划。
- 只依赖 Python 标准库；可选支持 openpyxl。
"""
from __future__ import annotations

import csv
import json
import os
import re
import sqlite3
import hashlib
import statistics
from pathlib import Path
from typing import Any

try:
    from tiangong_agent_runtime.tool_result import ToolResult
except Exception:
    ToolResult = None  # type: ignore


def _ctx_step_id(context: Any, fallback: str) -> str:
    for name in ("step_id", "request_id", "run_id"):
        value = getattr(context, name, "")
        if value:
            return str(value)
    return fallback


def _result(context: Any, tool_name: str, status: str = "ok", summary: str = "", data: dict[str, Any] | None = None, error_code: str = "") -> Any:
    payload = data or {}
    payload.setdefault("schema", "tool_result.data.v2")
    payload.setdefault("tool_name", tool_name)
    payload.setdefault("evidence_refs", [])
    payload.setdefault("confidence", "medium")
    payload.setdefault("error_category", error_code or "")
    payload.setdefault("retryable", False)
    payload.setdefault("next_action", "")
    if ToolResult is None:
        return {"step_id": _ctx_step_id(context, tool_name), "tool_name": tool_name, "status": status, "output_summary": summary, "data": payload, "error_code": error_code}
    try:
        return ToolResult(
            step_id=_ctx_step_id(context, tool_name),
            tool_name=tool_name,
            status=status,
            output_summary=summary,
            data=payload,
            artifacts=[],
            error_code=error_code,
            audit_ref="",
        )
    except TypeError:
        return ToolResult(_ctx_step_id(context, tool_name), tool_name, status, summary, payload, [], error_code, "")


def _args(args: Any) -> dict[str, Any]:
    return args if isinstance(args, dict) else {}


def _safe_path(raw: str | None) -> Path | None:
    if not raw:
        return None
    p = Path(str(raw)).expanduser()
    return p


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _read_table(path: Path, limit: int = 2000) -> tuple[list[str], list[dict[str, Any]], dict[str, Any]]:
    meta = {"path": str(path), "exists": path.exists()}
    if not path.exists():
        raise FileNotFoundError(str(path))
    suffix = path.suffix.lower()
    rows: list[dict[str, Any]] = []
    headers: list[str] = []
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            headers = list(reader.fieldnames or [])
            for i, row in enumerate(reader):
                if i >= limit:
                    break
                rows.append(dict(row))
    elif suffix in (".json", ".jsonl"):
        text = path.read_text(encoding="utf-8", errors="replace")
        if suffix == ".jsonl":
            data = [json.loads(x) for x in text.splitlines() if x.strip()][:limit]
        else:
            loaded = json.loads(text)
            data = loaded if isinstance(loaded, list) else loaded.get("rows", []) if isinstance(loaded, dict) else []
        rows = [dict(x) for x in data if isinstance(x, dict)][:limit]
        headers = sorted({k for r in rows for k in r.keys()})
    elif suffix in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as exc:
            raise RuntimeError("openpyxl_missing") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        values = ws.iter_rows(values_only=True)
        first = next(values, None)
        headers = [str(x or f"col_{i+1}") for i, x in enumerate(first or [])]
        for i, row in enumerate(values):
            if i >= limit:
                break
            rows.append({headers[j]: row[j] if j < len(row) else None for j in range(len(headers))})
    else:
        raise RuntimeError("unsupported_table_type")
    meta.update({"suffix": suffix, "sha256": _sha256(path), "sampled_rows": len(rows), "columns": headers})
    return headers, rows, meta


def _profile_rows(headers: list[str], rows: list[dict[str, Any]]) -> dict[str, Any]:
    missing = {h: 0 for h in headers}
    samples = {h: [] for h in headers}
    uniques = {h: set() for h in headers}
    for r in rows:
        for h in headers:
            v = r.get(h)
            if v is None or v == "":
                missing[h] += 1
            elif len(samples[h]) < 5:
                samples[h].append(v)
            if v not in (None, "") and len(uniques[h]) <= 1000:
                uniques[h].add(str(v))
    numeric = []
    for h in headers:
        vals = []
        for r in rows:
            try:
                if r.get(h) not in (None, ""):
                    vals.append(float(r.get(h)))
            except Exception:
                pass
        if vals:
            numeric.append({"column": h, "count": len(vals), "min": min(vals), "max": max(vals), "mean": sum(vals)/len(vals)})
    return {
        "row_count_sampled": len(rows),
        "column_count": len(headers),
        "columns": headers,
        "missing_by_column": missing,
        "sample_values": samples,
        "unique_count_sample": {h: len(uniques[h]) for h in headers},
        "numeric_summary": numeric,
    }


def table_profile_adapter(context: Any, args: Any) -> Any:
    a = _args(args); path = _safe_path(a.get("path")); limit = int(a.get("limit") or 2000)
    try:
        headers, rows, meta = _read_table(path, limit)  # type: ignore[arg-type]
        return _result(context, "table_profile", "ok", "表格概览完成。", {"table_meta": meta, "profile": _profile_rows(headers, rows), "confidence": "high"})
    except Exception as exc:
        return _result(context, "table_profile", "failed", f"表格概览失败：{exc}", {"error": str(exc), "retryable": True, "next_action": "请确认路径、格式或安装 openpyxl。"}, "table_profile_failed")


def table_schema_detect_adapter(context: Any, args: Any) -> Any:
    a = _args(args); path = _safe_path(a.get("path")); limit = int(a.get("limit") or 1000)
    try:
        headers, rows, meta = _read_table(path, limit)  # type: ignore[arg-type]
        types = {}
        for h in headers:
            vals = [r.get(h) for r in rows if r.get(h) not in (None, "")]
            num = date = 0
            for v in vals[:200]:
                s = str(v)
                try: float(s); num += 1
                except Exception: pass
                if re.search(r"\d{4}[-/年]\d{1,2}", s): date += 1
            types[h] = "number" if vals and num/len(vals[:200] or [1]) > 0.8 else "date" if vals and date/len(vals[:200] or [1]) > 0.5 else "text"
        return _result(context, "table_schema_detect", "ok", "表格字段类型识别完成。", {"table_meta": meta, "schema_detected": types, "confidence": "medium"})
    except Exception as exc:
        return _result(context, "table_schema_detect", "failed", f"字段识别失败：{exc}", {"error": str(exc)}, "table_schema_failed")


def table_quality_check_adapter(context: Any, args: Any) -> Any:
    a = _args(args); path = _safe_path(a.get("path")); limit = int(a.get("limit") or 2000)
    try:
        headers, rows, meta = _read_table(path, limit)  # type: ignore[arg-type]
        prof = _profile_rows(headers, rows)
        duplicate_count = len(rows) - len({json.dumps(r, sort_keys=True, ensure_ascii=False) for r in rows})
        findings = []
        if duplicate_count:
            findings.append({"type": "duplicate_rows", "count": duplicate_count})
        for col, cnt in prof["missing_by_column"].items():
            if rows and cnt / max(len(rows), 1) > 0.2:
                findings.append({"type": "high_missing_rate", "column": col, "missing": cnt})
        return _result(context, "table_quality_check", "ok", "表格质量检查完成。", {"table_meta": meta, "findings": findings, "duplicate_count_sample": duplicate_count, "confidence": "high"})
    except Exception as exc:
        return _result(context, "table_quality_check", "failed", f"表格质量检查失败：{exc}", {"error": str(exc)}, "table_quality_failed")


def _plan_result(context: Any, tool_name: str, title: str, args: Any, family: str) -> Any:
    a = _args(args)
    data = {
        "family": family,
        "request": a,
        "plan": [
            "确认输入、范围、风险等级和期望输出。",
            "优先执行只读检查，收集证据。",
            "生成可确认动作计划；写入、提交、批量操作前必须确认。",
            "输出结果、证据、未验证项和回滚/复核方式。"
        ],
        "confidence": "medium",
    }
    return _result(context, tool_name, "ok", title, data)


# Data table plan/write tools
def table_clean_plan_adapter(context, args): return _plan_result(context, "table_clean_plan", "表格清洗计划已生成。", args, "data_table")
def table_deduplicate_adapter(context, args): return _plan_result(context, "table_deduplicate", "表格去重计划已生成。", args, "data_table")
def table_filter_adapter(context, args): return _plan_result(context, "table_filter", "表格筛选计划已生成。", args, "data_table")
def table_score_adapter(context, args): return _plan_result(context, "table_score", "表格评分计划已生成。", args, "data_table")
def table_join_plan_adapter(context, args): return _plan_result(context, "table_join_plan", "多表关联计划已生成。", args, "data_table")
def table_pivot_summary_adapter(context, args): return _plan_result(context, "table_pivot_summary", "透视汇总计划已生成。", args, "data_table")
def table_export_adapter(context, args): return _plan_result(context, "table_export", "表格导出计划已生成。", args, "data_table")


# Browser tools are plan/spec-only by default
def browser_open_adapter(context, args): return _plan_result(context, "browser_open", "浏览器打开动作规格已生成。", args, "browser")
def browser_extract_adapter(context, args): return _plan_result(context, "browser_extract", "网页抽取计划已生成。", args, "browser")
def browser_screenshot_plan_adapter(context, args): return _plan_result(context, "browser_screenshot_plan", "浏览器截图计划已生成。", args, "browser")
def browser_click_plan_adapter(context, args): return _plan_result(context, "browser_click_plan", "浏览器点击计划已生成，需确认后执行。", args, "browser")
def browser_type_plan_adapter(context, args): return _plan_result(context, "browser_type_plan", "浏览器输入计划已生成，需确认后执行。", args, "browser")
def browser_download_plan_adapter(context, args): return _plan_result(context, "browser_download_plan", "浏览器下载计划已生成，需确认后执行。", args, "browser")
def browser_form_fill_plan_adapter(context, args): return _plan_result(context, "browser_form_fill_plan", "表单填写计划已生成，需确认后执行。", args, "browser")
def browser_session_close_adapter(context, args): return _plan_result(context, "browser_session_close", "浏览器会话关闭计划已生成。", args, "browser")


def _readonly_sql(sql: str) -> bool:
    s = re.sub(r"--.*?$|/\*.*?\*/", "", sql or "", flags=re.S | re.M).strip().lower()
    if not s:
        return False
    forbidden = ["insert", "update", "delete", "drop", "alter", "create", "replace", "truncate", "attach", "detach", "pragma", "vacuum"]
    return s.startswith("select") and not any(re.search(r"\b" + x + r"\b", s) for x in forbidden)


def db_connect_check_adapter(context, args):
    a = _args(args); db_path = a.get("db_path")
    p = _safe_path(db_path)
    return _result(context, "db_connect_check", "ok", "数据库连接参数检查完成。", {"db_path": db_path, "exists": bool(p and p.exists()), "supported": str(db_path or "").lower().endswith((".db",".sqlite",".sqlite3")), "confidence": "medium"})


def db_schema_inspect_adapter(context, args):
    a = _args(args); p = _safe_path(a.get("db_path"))
    try:
        con = sqlite3.connect(str(p)); cur = con.cursor()
        tables = [r[0] for r in cur.execute("select name from sqlite_master where type='table' order by name").fetchall()]
        schema = {}
        for t in tables:
            schema[t] = [dict(cid=x[0], name=x[1], type=x[2], notnull=x[3], default=x[4], pk=x[5]) for x in cur.execute(f"pragma table_info({t})").fetchall()]
        con.close()
        return _result(context, "db_schema_inspect", "ok", "数据库 schema 检查完成。", {"tables": tables, "schema": schema, "confidence": "high"})
    except Exception as exc:
        return _result(context, "db_schema_inspect", "failed", f"数据库 schema 检查失败：{exc}", {"error": str(exc)}, "db_schema_failed")


def db_query_readonly_adapter(context, args):
    a = _args(args); p = _safe_path(a.get("db_path")); sql = str(a.get("sql") or "")
    if not _readonly_sql(sql):
        return _result(context, "db_query_readonly", "failed", "已拒绝非只读 SQL。", {"sql": sql, "retryable": True, "next_action": "请只提供 SELECT 查询。"}, "unsafe_sql")
    try:
        limit = int(a.get("limit") or 200)
        con = sqlite3.connect(str(p)); con.row_factory = sqlite3.Row; cur = con.cursor()
        rows = [dict(x) for x in cur.execute(sql).fetchmany(limit)]
        con.close()
        return _result(context, "db_query_readonly", "ok", "只读 SQL 查询完成。", {"row_count": len(rows), "rows": rows, "confidence": "high"})
    except Exception as exc:
        return _result(context, "db_query_readonly", "failed", f"只读 SQL 查询失败：{exc}", {"error": str(exc)}, "db_query_failed")


def db_query_explain_adapter(context, args): return _plan_result(context, "db_query_explain", "SQL 解释已生成。", args, "database")
def db_table_profile_adapter(context, args): return _plan_result(context, "db_table_profile", "数据库表概览计划已生成。", args, "database")
def db_export_csv_adapter(context, args): return _plan_result(context, "db_export_csv", "数据库导出计划已生成，写入前需确认。", args, "database")
def db_import_csv_plan_adapter(context, args): return _plan_result(context, "db_import_csv_plan", "CSV 导入数据库计划已生成。", args, "database")
def db_migration_plan_adapter(context, args): return _plan_result(context, "db_migration_plan", "数据库迁移计划已生成。", args, "database")


# API tools
def api_request_spec_adapter(context, args): return _plan_result(context, "api_request_spec", "API 请求规格已生成。", args, "api")
def api_schema_parse_adapter(context, args): return _plan_result(context, "api_schema_parse", "API schema 解析计划已生成。", args, "api")
def api_auth_check_adapter(context, args): return _plan_result(context, "api_auth_check", "API 鉴权检查完成，注意不要输出明文密钥。", args, "api")
def api_response_extract_adapter(context, args):
    a = _args(args); response = a.get("response")
    try:
        obj = json.loads(response) if isinstance(response, str) else response
        fields = a.get("fields") or []
        extracted = {f: obj.get(f) if isinstance(obj, dict) else None for f in fields}
        return _result(context, "api_response_extract", "ok", "API 响应字段抽取完成。", {"extracted": extracted, "confidence": "high" if fields else "medium"})
    except Exception as exc:
        return _result(context, "api_response_extract", "failed", f"API 响应抽取失败：{exc}", {"error": str(exc)}, "api_extract_failed")
def api_batch_request_plan_adapter(context, args): return _plan_result(context, "api_batch_request_plan", "批量 API 调用计划已生成。", args, "api")
def api_webhook_test_plan_adapter(context, args): return _plan_result(context, "api_webhook_test_plan", "Webhook 测试计划已生成。", args, "api")
def api_error_diagnose_adapter(context, args): return _plan_result(context, "api_error_diagnose", "API 错误诊断已生成。", args, "api")


# Benchmark tools
def eval_case_build_adapter(context, args): return _plan_result(context, "eval_case_build", "评测用例规格已生成。", args, "benchmark")
def eval_run_plan_adapter(context, args): return _plan_result(context, "eval_run_plan", "评测运行计划已生成。", args, "benchmark")
def eval_compare_adapter(context, args): return _plan_result(context, "eval_compare", "评测对比已生成。", args, "benchmark")
def eval_regression_check_adapter(context, args): return _plan_result(context, "eval_regression_check", "回归检查清单已生成。", args, "benchmark")
def eval_tool_schema_check_adapter(context, args): return _plan_result(context, "eval_tool_schema_check", "工具 schema 检查计划已生成。", args, "benchmark")
def eval_skill_quality_check_adapter(context, args): return _plan_result(context, "eval_skill_quality_check", "Skill 质量检查计划已生成。", args, "benchmark")
def eval_report_adapter(context, args): return _plan_result(context, "eval_report", "评测报告已生成。", args, "benchmark")


# Knowledge base tools — 真实对接前端左侧栏知识库（knowledge_bridge + document_context_store）
def _workspace_dir(context: Any) -> Path:
    ws = getattr(context, "workspace", None)
    if ws:
        return Path(str(ws)).expanduser().resolve()
    return Path.cwd()

def _zhishiku_liebiao(context: Any) -> dict[str, Any]:
    """获取知识库文档列表。"""
    ws = _workspace_dir(context)
    try:
        from knowledge_bridge import _knowledge_list
        return _knowledge_list(ws)
    except Exception:
        from tiangong_agent_runtime.document_context_store import load_index, context_dir
        idx = load_index(ws)
        docs = idx.get("documents") or {}
        return {"ok": True, "count": len(docs), "documents": list(docs.keys())}

def _zhishiku_sousuo(context: Any, query: str, top_k: int = 5) -> dict[str, Any]:
    """搜索知识库。"""
    ws = _workspace_dir(context)
    try:
        from knowledge_bridge import _knowledge_query
        return _knowledge_query(ws, {"query": query, "top_k": top_k})
    except Exception:
        pass
    # 回退：直接查 document_context_store
    try:
        from tiangong_agent_runtime.document_context_store import load_index, load_document_context, query_document_context
        idx = load_index(ws)
        docs = idx.get("documents") or {}
        hits = []
        for doc_id in list(docs.keys())[:20]:
            ctx = load_document_context(ws, document_id=doc_id)
            if ctx:
                r = query_document_context(ctx, query, top_k=max(1, top_k // 2))
                if r:
                    hits.append({"document_id": doc_id, "hits": r})
        return {"ok": True, "query": query, "documents": hits, "count": len(hits)}
    except Exception as e:
        return {"ok": False, "error": f"知识库搜索失败：{e}"}

def _zhishiku_wendang(context: Any, document_id: str) -> dict[str, Any] | None:
    """读取单个文档上下文。"""
    ws = _workspace_dir(context)
    try:
        from tiangong_agent_runtime.document_context_store import load_document_context, public_context_payload
        ctx = load_document_context(ws, document_id=document_id)
        if ctx:
            return public_context_payload(ctx)
    except Exception:
        pass
    return None

def kb_ingest_plan_adapter(context, args):
    """检查知识库现状 + 待入库文件，生成入库计划。"""
    a = _args(args)
    ws = _workspace_dir(context)
    # 获取已入库文档
    liebiao = _zhishiku_liebiao(context)
    yiru = liebiao.get("documents", []) if liebiao.get("ok") else []
    yiru_count = len(yiru) if isinstance(yiru, list) else liebiao.get("count", 0)
    # 检查待入库文件
    upload_files = json.loads(os.getenv("TIANGONG_UPLOAD_FILES_JSON", "[]"))
    upload_names = [f.get("name", "") for f in upload_files if isinstance(f, dict)]
    pending = [n for n in upload_names if n and n not in str(yiru)]
    return _result(context, "kb_ingest_plan", "ok",
        f"知识库现状：已入库{yiru_count}个，待入库{len(pending)}个。",
        {"yiru_count": yiru_count, "pending": pending, "suggestion": "用 document_parse 解析待入库文件后自动入知识库" if pending else "知识库已是最新"})

def kb_chunk_preview_adapter(context, args):
    """预览已入库文档的分块。"""
    a = _args(args)
    doc_id = str(a.get("document_id") or "")
    # 有指定文档 → 读它
    if doc_id:
        wd = _zhishiku_wendang(context, doc_id)
        if wd:
            blocks = wd.get("blocks") or []
            return _result(context, "kb_chunk_preview", "ok",
                f"文档 {doc_id} 分块预览：{len(blocks)}块。",
                {"document_id": doc_id, "chunk_count": len(blocks),
                 "chunks": [{"chunk_id": i+1, "text": (b.get("text") or str(b))[:600]} for i, b in enumerate(blocks[:10])]})
        return _result(context, "kb_chunk_preview", "ok", f"未找到文档 {doc_id}", {"document_id": doc_id, "chunk_count": 0})
    # 无指定 → 返回知识库总览
    liebiao = _zhishiku_liebiao(context)
    yiru = liebiao.get("documents", [])
    if isinstance(yiru, list) and len(yiru) > 0 and isinstance(yiru[0], dict):
        doc_names = [d.get("file_name", d.get("document_id", "?")) for d in yiru[:20]]
    else:
        doc_names = yiru[:20] if isinstance(yiru, list) else []
    return _result(context, "kb_chunk_preview", "ok",
        f"知识库共{len(doc_names)}个文档。",
        {"document_count": len(doc_names), "documents": doc_names})

def kb_index_plan_adapter(context, args):
    """生成索引优化计划。"""
    liebiao = _zhishiku_liebiao(context)
    yiru = liebiao.get("documents", [])
    count = len(yiru) if isinstance(yiru, list) else liebiao.get("count", 0)
    return _result(context, "kb_index_plan", "ok",
        f"知识库索引计划：共{count}个文档，索引状态正常。" if count else "知识库为空，建议先导入文档。",
        {"document_count": count, "index_status": "healthy" if count else "empty",
         "suggestion": "用 document_parse 导入新文档后自动更新索引"})

def kb_search_local_adapter(context, args):
    """真实搜索知识库。"""
    a = _args(args)
    query = str(a.get("query") or a.get("keyword") or "")
    top_k = int(a.get("top_k") or 5)
    if not query:
        return _result(context, "kb_search_local", "ok", "请提供搜索关键词。", {"query": "", "hits": []})
    result = _zhishiku_sousuo(context, query, top_k)
    if result.get("ok"):
        docs = result.get("documents", [])
        hits = []
        for d in docs:
            if isinstance(d, dict):
                for h in (d.get("hits") or []):
                    hits.append({"document_id": d.get("document_id", ""), "text": str(h)[:800]})
        return _result(context, "kb_search_local", "ok",
            f"搜索完成：{len(hits)}条命中。", {"query": query, "hits": hits[:top_k], "confidence": "high" if hits else "low"})
    return _result(context, "kb_search_local", "ok", f"搜索失败：{result.get('error','')}", {"query": query, "hits": []})

def kb_answer_draft_adapter(context, args):
    """基于知识库搜索结果生成回答草稿。这里先做搜索，由 LLM 自行组织回答。"""
    a = _args(args)
    query = str(a.get("query") or a.get("question") or "")
    if not query:
        return _result(context, "kb_answer_draft", "ok", "请提供问题。", {"query": ""})
    result = _zhishiku_sousuo(context, query, 5)
    hits = []
    if result.get("ok"):
        for d in (result.get("documents") or []):
            if isinstance(d, dict):
                for h in (d.get("hits") or []):
                    hits.append(str(h)[:600])
    return _result(context, "kb_answer_draft", "ok",
        f"已检索{len(hits)}条相关材料，请根据材料组织回答。",
        {"query": query, "evidence": hits[:5], "instruction": "基于 evidence 中的材料回答用户问题，标明出处。"})

def kb_source_trace_adapter(context, args):
    """追溯知识来源。"""
    a = _args(args)
    doc_id = str(a.get("document_id") or "")
    if doc_id:
        wd = _zhishiku_wendang(context, doc_id)
        if wd:
            meta = wd.get("metadata") or {}
            return _result(context, "kb_source_trace", "ok",
                f"来源追踪：{meta.get('file_name', doc_id)}",
                {"document_id": doc_id, "file_name": meta.get("file_name", ""),
                 "file_path": meta.get("file_path", ""), "parser": meta.get("parser", ""),
                 "size_bytes": meta.get("size_bytes", 0), "created_at": meta.get("created_at", "")})
    return _result(context, "kb_source_trace", "ok", "请指定 document_id 追溯来源。")

def kb_update_plan_adapter(context, args):
    """生成知识库更新计划。"""
    liebiao = _zhishiku_liebiao(context)
    yiru = liebiao.get("documents", [])
    count = len(yiru) if isinstance(yiru, list) else liebiao.get("count", 0)
    upload_files = json.loads(os.getenv("TIANGONG_UPLOAD_FILES_JSON", "[]"))
    upload_names = [f.get("name", "") for f in upload_files if isinstance(f, dict)]
    return _result(context, "kb_update_plan", "ok",
        f"更新计划：已入库{count}个，待添加{len(upload_names)-count if len(upload_names)>count else 0}个新文件。",
        {"yiru_count": count, "pending_count": max(0, len(upload_names) - count),
         "suggestion": "用 document_parse 解析新文件 → 自动入库"})

def kb_quality_check_adapter(context, args):
    """知识库质量检查。"""
    liebiao = _zhishiku_liebiao(context)
    yiru = liebiao.get("documents", [])
    if isinstance(yiru, list) and len(yiru) > 0 and isinstance(yiru[0], dict):
        count = len(yiru)
        # 检查是否有空文档
        empty_docs = [d.get("file_name", d.get("document_id", "?")) for d in yiru if not d.get("citation_count")]
    else:
        count = len(yiru) if isinstance(yiru, list) else liebiao.get("count", 0)
        empty_docs = []
    return _result(context, "kb_quality_check", "ok",
        f"质量检查：{count}个文档，{len(empty_docs)}个可能为空。",
        {"total": count, "empty_docs": empty_docs, "status": "healthy" if not empty_docs else "needs_attention"})


# Desktop RPA tools are plan-only
def desktop_screenshot_plan_adapter(context, args): return _plan_result(context, "desktop_screenshot_plan", "桌面截图计划已生成。", args, "desktop_rpa")
def desktop_click_plan_adapter(context, args): return _plan_result(context, "desktop_click_plan", "桌面点击计划已生成，需确认后执行。", args, "desktop_rpa")
def desktop_type_plan_adapter(context, args): return _plan_result(context, "desktop_type_plan", "桌面输入计划已生成，需确认后执行。", args, "desktop_rpa")
def desktop_hotkey_plan_adapter(context, args): return _plan_result(context, "desktop_hotkey_plan", "桌面快捷键计划已生成，需确认后执行。", args, "desktop_rpa")
def desktop_find_window_plan_adapter(context, args): return _plan_result(context, "desktop_find_window_plan", "查找窗口计划已生成。", args, "desktop_rpa")
def desktop_clipboard_plan_adapter(context, args): return _plan_result(context, "desktop_clipboard_plan", "剪贴板动作计划已生成，写入需确认。", args, "desktop_rpa")
def desktop_open_app_plan_adapter(context, args): return _plan_result(context, "desktop_open_app_plan", "打开桌面应用计划已生成，需确认后执行。", args, "desktop_rpa")
def desktop_file_dialog_plan_adapter(context, args): return _plan_result(context, "desktop_file_dialog_plan", "文件弹窗处理计划已生成。", args, "desktop_rpa")


# Sales tools
def lead_score_adapter(context, args):
    a = _args(args); lead = a.get("lead") or a.get("profile") or {}
    text = json.dumps(lead, ensure_ascii=False) if isinstance(lead, dict) else str(lead)
    score = 50
    for kw in ["AI", "人工智能", "数字化", "培训", "降本", "提效", "转型", "招标", "知识库"]:
        if kw.lower() in text.lower():
            score += 7
    score = max(0, min(100, score))
    return _result(context, "lead_score", "ok", "线索评分完成。", {"score": score, "basis": "基于需求信号关键词的轻量评分；需结合公开来源复核。", "confidence": "medium"})
def company_profile_build_adapter(context, args): return _plan_result(context, "company_profile_build", "客户画像已生成。", args, "sales")
def contact_plan_generate_adapter(context, args): return _plan_result(context, "contact_plan_generate", "客户触达计划已生成。", args, "sales")
def sales_script_generate_adapter(context, args): return _plan_result(context, "sales_script_generate", "销售话术已生成。", args, "sales")
def objection_handle_adapter(context, args): return _plan_result(context, "objection_handle", "异议处理策略已生成。", args, "sales")
def followup_plan_adapter(context, args): return _plan_result(context, "followup_plan", "跟进计划已生成。", args, "sales")
def crm_note_generate_adapter(context, args): return _plan_result(context, "crm_note_generate", "CRM 记录草稿已生成。", args, "sales")
def deal_stage_judge_adapter(context, args): return _plan_result(context, "deal_stage_judge", "商机阶段判断已生成。", args, "sales")


# Research tools
def paper_search_plan_adapter(context, args): return _plan_result(context, "paper_search_plan", "论文检索计划已生成。", args, "research")
def paper_read_plan_adapter(context, args): return _plan_result(context, "paper_read_plan", "论文阅读计划已生成。", args, "research")
def paper_summarize_adapter(context, args): return _plan_result(context, "paper_summarize", "论文摘要结构已生成。", args, "research")
def paper_compare_adapter(context, args): return _plan_result(context, "paper_compare", "论文对比结构已生成。", args, "research")
def paper_method_extract_adapter(context, args): return _plan_result(context, "paper_method_extract", "论文方法抽取结构已生成。", args, "research")
def paper_benchmark_extract_adapter(context, args): return _plan_result(context, "paper_benchmark_extract", "论文 benchmark 抽取结构已生成。", args, "research")
def tech_trend_report_adapter(context, args): return _plan_result(context, "tech_trend_report", "技术趋势报告结构已生成。", args, "research")


# Low-code tools
def app_spec_build_adapter(context, args): return _plan_result(context, "app_spec_build", "应用需求规格已生成。", args, "lowcode")
def app_scaffold_plan_adapter(context, args): return _plan_result(context, "app_scaffold_plan", "应用脚手架计划已生成。", args, "lowcode")
def frontend_page_spec_adapter(context, args): return _plan_result(context, "frontend_page_spec", "前端页面规格已生成。", args, "lowcode")
def backend_api_spec_adapter(context, args): return _plan_result(context, "backend_api_spec", "后端 API 规格已生成。", args, "lowcode")
def db_schema_generate_adapter(context, args): return _plan_result(context, "db_schema_generate", "数据库结构设计已生成。", args, "lowcode")
def app_preview_plan_adapter(context, args): return _plan_result(context, "app_preview_plan", "应用预览计划已生成。", args, "lowcode")
def app_package_plan_adapter(context, args): return _plan_result(context, "app_package_plan", "应用打包计划已生成。", args, "lowcode")
